"""
Utilidades de exportacion a Excel y JSON.

Encapsulan la logica de construccion de los archivos de salida para que sea
reutilizable tanto desde los endpoints REST como desde scripts standalone.
"""
from __future__ import annotations

import io
import json
from typing import Dict, List, Optional


def build_excel(
    circuit_info: dict,
    voltage_profile: Optional[List[dict]] = None,
    losses_elements: Optional[List[dict]] = None,
    losses_summary: Optional[dict] = None,
    hosting_capacity: Optional[List[dict]] = None,
    simulation_comparison: Optional[List[dict]] = None,
    violations: Optional[dict] = None,
) -> bytes:
    """
    Construye un archivo Excel con multiples hojas y retorna los bytes.
    Requiere openpyxl instalado.
    """
    import openpyxl

    wb = openpyxl.Workbook()

    # --- Hoja: Circuito ---
    ws_circuit = wb.active
    ws_circuit.title = "Circuito"
    ws_circuit.append(["Campo", "Valor"])
    for k, v in circuit_info.items():
        ws_circuit.append([k, str(v)])

    # --- Hoja: Voltajes_Base ---
    if voltage_profile:
        ws_v = wb.create_sheet("Voltajes_Base")
        ws_v.append(["Bus_Fase", "Voltaje PU", "En rango (0.95-1.05)"])
        for row in voltage_profile:
            ws_v.append([row["bus_phase"], row["voltage_pu"], row["in_range"]])

    # --- Hoja: Perdidas_Base ---
    if losses_elements:
        ws_l = wb.create_sheet("Perdidas_Base")
        if losses_summary:
            ws_l.append(["=== Resumen ==="])
            for k, v in losses_summary.items():
                ws_l.append([k, v])
            ws_l.append([])
        ws_l.append(["Tipo", "Elemento", "kW Perdida", "kvar Perdida", "% Potencia"])
        for el in losses_elements:
            ws_l.append(
                [el["type"], el["element"], el["losses_kw"], el["losses_kvar"], el["losses_pct"]]
            )

    # --- Hoja: Hosting_Capacity ---
    if hosting_capacity:
        ws_hc = wb.create_sheet("Hosting_Capacity")
        ws_hc.append(["Barra", "Fase", "Max GD sin violacion (kW)", "Restriccion limitante"])
        for entry in hosting_capacity:
            ws_hc.append(
                [
                    entry.get("bus"),
                    entry.get("phase"),
                    entry.get("max_gd_kw"),
                    entry.get("limiting_constraint"),
                ]
            )

    # --- Hoja: Simulacion_Voltajes ---
    if simulation_comparison:
        ws_sv = wb.create_sheet("Simulacion_Voltajes")
        ws_sv.append(
            [
                "Bus_Fase",
                "Voltaje Base PU",
                "Voltaje con GD PU",
                "Delta PU",
                "En rango base",
                "En rango con GD",
            ]
        )
        for row in simulation_comparison:
            ws_sv.append(
                [
                    row.get("bus_phase"),
                    row.get("voltage_base_pu"),
                    row.get("voltage_with_gd_pu"),
                    row.get("delta_pu"),
                    row.get("in_range_base"),
                    row.get("in_range_with_gd"),
                ]
            )

    # --- Hoja: Violaciones ---
    if violations:
        ws_viol = wb.create_sheet("Violaciones")
        all_violations = []
        for tipo, vlist in violations.items():
            for v in vlist:
                all_violations.append({"tipo": tipo, **v})
        if all_violations:
            headers = list(all_violations[0].keys())
            ws_viol.append(headers)
            for v in all_violations:
                ws_viol.append([v.get(h) for h in headers])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def build_json_export(
    circuit_id: str,
    circuit_info: dict,
    voltage_profile: Optional[List[dict]] = None,
    losses_summary: Optional[dict] = None,
    losses_elements: Optional[List[dict]] = None,
    hosting_capacity: Optional[List[dict]] = None,
    simulations: Optional[List[dict]] = None,
    exported_at: Optional[str] = None,
) -> dict:
    """
    Construye el payload JSON de exportacion completo.
    """
    import datetime

    return {
        "circuit_id": circuit_id,
        "exported_at": exported_at
        or (datetime.datetime.utcnow().isoformat() + "Z"),
        "circuit_info": circuit_info,
        "voltage_profile": voltage_profile or [],
        "losses": {
            "summary": losses_summary or {},
            "elements": losses_elements or [],
        },
        "hosting_capacity": hosting_capacity,
        "simulations": simulations or [],
    }
