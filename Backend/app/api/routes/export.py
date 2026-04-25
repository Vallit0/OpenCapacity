"""
Grupo 6 — Exportacion.

GET /{circuit_id}/export/excel  Exporta resultados a Excel (.xlsx).
GET /{circuit_id}/export/json   Exporta todos los resultados a JSON.
"""
from __future__ import annotations

import datetime
import io
import json

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse

from app.api.dependencies import get_redis, require_circuit, dss_lock
from app.core.dss_engine import DSSEngine, DSSEngineError
from app.core.logging_config import get_logger, log_timer

logger = get_logger(__name__)
router = APIRouter()


@router.get("/{circuit_id}/export/excel")
async def export_excel(
    circuit_id: str,
    include_voltage_profile: bool = Query(True),
    include_losses: bool = Query(True),
    include_hosting_capacity: bool = Query(True),
    include_violations: bool = Query(True),
    simulation_id: str = Query(None),
):
    """
    Exporta los resultados disponibles a un archivo Excel con multiples hojas.
    Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    """
    logger.info(
        "EXPORT excel | circuit_id=%s | voltaje=%s | perdidas=%s | hosting=%s",
        circuit_id, include_voltage_profile, include_losses, include_hosting_capacity,
    )

    try:
        import openpyxl
    except ImportError:
        logger.error("EXPORT excel FALLIDA | openpyxl no instalado")
        raise HTTPException(
            status_code=500,
            detail="openpyxl no esta instalado. Agregue 'openpyxl' a requirements.txt.",
        )

    r = get_redis()
    circuit_data = require_circuit(circuit_id, r)

    info_raw = r.get(f"circuit:{circuit_id}:info")
    circuit_info = json.loads(info_raw) if info_raw else {}
    circuit_name = circuit_info.get("name", circuit_id)

    try:
        async with dss_lock:
            with log_timer(logger, "export_excel_dss", circuit_id=circuit_id):
                engine = DSSEngine()
                engine.load_circuit(
                    circuit_data["dss_content"], circuit_data["linecodes_content"]
                )
                voltage_profile = engine.get_voltage_profile() if include_voltage_profile else []
                elements, losses_summary = engine.get_losses() if include_losses else ([], {})
    except DSSEngineError as exc:
        logger.error("EXPORT excel FALLIDA | circuit_id=%s | DSSEngineError | %s", circuit_id, exc)
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    wb = openpyxl.Workbook()
    sheets_created = []

    ws_circuit = wb.active
    ws_circuit.title = "Circuito"
    ws_circuit.append(["Campo", "Valor"])
    for k, v in circuit_info.items():
        ws_circuit.append([k, v])
    sheets_created.append("Circuito")

    if include_voltage_profile and voltage_profile:
        ws_v = wb.create_sheet("Voltajes_Base")
        ws_v.append(["Bus_Fase", "Voltaje PU", "En rango (0.95-1.05)"])
        for row in voltage_profile:
            ws_v.append([row["bus_phase"], row["voltage_pu"], row["in_range"]])
        sheets_created.append("Voltajes_Base")

    if include_losses and elements:
        ws_l = wb.create_sheet("Perdidas_Base")
        ws_l.append(["Tipo", "Elemento", "kW Perdida", "kvar Perdida", "% Potencia"])
        for el in elements:
            ws_l.append(
                [el["type"], el["element"], el["losses_kw"], el["losses_kvar"], el["losses_pct"]]
            )
        sheets_created.append("Perdidas_Base")

    if include_hosting_capacity:
        results_raw = r.get(f"hosting_capacity:{circuit_id}:results")
        if results_raw:
            results = json.loads(results_raw)
            ws_hc = wb.create_sheet("Hosting_Capacity")
            ws_hc.append(["Barra", "Fase", "Max GD sin violacion (kW)", "Restriccion limitante"])
            for entry in results:
                ws_hc.append(
                    [
                        entry["bus"],
                        entry["phase"],
                        entry.get("max_gd_kw"),
                        entry.get("limiting_constraint"),
                    ]
                )
            sheets_created.append("Hosting_Capacity")
        else:
            logger.debug("EXPORT excel | circuit_id=%s | sin datos de hosting capacity", circuit_id)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_size_kb = round(len(buffer.getvalue()) / 1024, 1)

    logger.info(
        "EXPORT excel OK | circuit_id=%s | hojas=%s | size=%.1f KB",
        circuit_id, sheets_created, file_size_kb,
    )

    filename = f"hosting_capacity_{circuit_name}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{circuit_id}/export/json")
async def export_json(circuit_id: str):
    """
    Exporta todos los resultados disponibles del circuito en formato JSON.
    """
    logger.info("EXPORT json | circuit_id=%s", circuit_id)

    r = get_redis()
    circuit_data = require_circuit(circuit_id, r)

    info_raw = r.get(f"circuit:{circuit_id}:info")
    circuit_info = json.loads(info_raw) if info_raw else {}

    try:
        async with dss_lock:
            with log_timer(logger, "export_json_dss", circuit_id=circuit_id):
                engine = DSSEngine()
                engine.load_circuit(
                    circuit_data["dss_content"], circuit_data["linecodes_content"]
                )
                voltage_profile = engine.get_voltage_profile()
                elements, losses_summary = engine.get_losses()
    except DSSEngineError as exc:
        logger.error("EXPORT json FALLIDA | circuit_id=%s | DSSEngineError | %s", circuit_id, exc)
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    results_raw = r.get(f"hosting_capacity:{circuit_id}:results")
    hosting_capacity = json.loads(results_raw) if results_raw else None

    exported_at = datetime.datetime.utcnow().isoformat() + "Z"

    logger.info(
        "EXPORT json OK | circuit_id=%s | voltajes=%d | elementos=%d | hosting=%s",
        circuit_id, len(voltage_profile), len(elements), hosting_capacity is not None,
    )

    payload = {
        "circuit_id": circuit_id,
        "exported_at": exported_at,
        "circuit_info": circuit_info,
        "voltage_profile": voltage_profile,
        "losses": {
            "summary": losses_summary,
            "elements": elements,
        },
        "hosting_capacity": hosting_capacity,
    }

    return JSONResponse(content=payload)
